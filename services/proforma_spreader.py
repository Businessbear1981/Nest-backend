"""Proforma Spreader — reads client Excel proformas and extracts financial assumptions.

Converts uploaded .xlsx into structured data matching credit_engine.py inputs.
Auto-detects: NOI row, revenue rows, expense rows, debt service, occupancy ramp.
Benchmarks assumptions against NEST market data.
"""
import os
from datetime import datetime

try:
    import openpyxl
except ImportError:
    openpyxl = None


# ── Senior Living Benchmarks (NIC MAP Q4 2025) ─────────────────

BENCHMARKS = {
    "senior_living": {
        "revenue_per_unit_monthly": {"il": (4800, 6500), "al": (5500, 8500), "mc": (6500, 10000)},
        "opex_per_unit_annual": {"il": (42000, 65000), "al": (68000, 95000), "mc": (88000, 120000)},
        "occupancy_ramp": {
            12: (0.15, 0.40),   # month 12: 15-40%
            24: (0.40, 0.72),   # month 24: 40-72%
            36: (0.72, 0.88),   # month 36: 72-88%
            48: (0.88, 0.95),   # month 48: 88-95%
        },
        "stabilization_months": (24, 42),
        "exit_cap_rate": (5.5, 8.25),
        "staffing_pct_of_revenue": (0.52, 0.65),
    },
    "industrial": {
        "rent_per_sf_annual": (8.50, 22.00),
        "opex_per_sf_annual": (1.50, 3.50),
        "vacancy_rate": (0.03, 0.09),
        "exit_cap_rate": (4.5, 7.25),
    },
    "multifamily": {
        "revenue_per_unit_monthly": (1200, 3200),
        "opex_per_unit_annual": (6500, 13000),
        "vacancy_rate": (0.04, 0.10),
        "exit_cap_rate": (4.25, 6.75),
    },
}

# Keywords for auto-detecting rows
REVENUE_KEYWORDS = ["revenue", "income", "rent", "care fee", "ancillary", "gross"]
EXPENSE_KEYWORDS = ["expense", "operating", "payroll", "staffing", "utilities", "insurance",
                     "maintenance", "admin", "management fee", "g&a", "general"]
NOI_KEYWORDS = ["noi", "net operating income", "net income", "ebitda"]
DEBT_SERVICE_KEYWORDS = ["debt service", "mortgage", "loan payment", "interest", "principal"]
OCCUPANCY_KEYWORDS = ["occupancy", "census", "utilization", "occupied"]


class ProformaSpreader:
    """Reads client Excel proformas and extracts structured financial data."""

    def spread(self, file_path: str) -> dict:
        """Read .xlsx and extract all financial assumptions.

        Returns structured dict matching credit_engine.py inputs.
        """
        if not openpyxl:
            return {"error": "openpyxl not installed — run pip install openpyxl"}
        if not os.path.exists(file_path):
            return {"error": f"File not found: {file_path}"}

        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        structure = self.detect_sheet_structure(wb)
        extracted = self._extract_values(ws, structure)
        extracted["file_path"] = file_path
        extracted["extracted_at"] = datetime.utcnow().isoformat()
        return extracted

    def detect_sheet_structure(self, wb) -> dict:
        """Find revenue, expense, NOI, debt service, and occupancy rows."""
        ws = wb.active
        row_map = {
            "revenue_rows": [],
            "expense_rows": [],
            "noi_row": None,
            "debt_service_row": None,
            "occupancy_row": None,
            "header_row": 1,
            "data_start_col": 2,
        }

        for row in ws.iter_rows(min_row=1, max_row=min(100, ws.max_row), max_col=1, values_only=False):
            cell = row[0]
            if cell.value is None:
                continue
            label = str(cell.value).lower().strip()

            if any(kw in label for kw in NOI_KEYWORDS):
                row_map["noi_row"] = cell.row
            elif any(kw in label for kw in DEBT_SERVICE_KEYWORDS):
                row_map["debt_service_row"] = cell.row
            elif any(kw in label for kw in OCCUPANCY_KEYWORDS):
                row_map["occupancy_row"] = cell.row
            elif any(kw in label for kw in REVENUE_KEYWORDS):
                row_map["revenue_rows"].append(cell.row)
            elif any(kw in label for kw in EXPENSE_KEYWORDS):
                row_map["expense_rows"].append(cell.row)

        return row_map

    def _extract_values(self, ws, structure: dict) -> dict:
        """Extract numeric values from detected rows."""
        result = {
            "revenue_items": [],
            "expense_items": [],
            "noi": None,
            "debt_service": None,
            "occupancy_schedule": [],
            "total_revenue": 0,
            "total_expenses": 0,
        }

        max_col = min(ws.max_column, 50)

        # Extract revenue
        for row_num in structure.get("revenue_rows", []):
            label = str(ws.cell(row=row_num, column=1).value or "")
            values = []
            for col in range(2, max_col + 1):
                val = ws.cell(row=row_num, column=col).value
                if isinstance(val, (int, float)):
                    values.append(val)
            if values:
                result["revenue_items"].append({"label": label, "values": values})
                result["total_revenue"] += sum(values) / max(len(values), 1)

        # Extract expenses
        for row_num in structure.get("expense_rows", []):
            label = str(ws.cell(row=row_num, column=1).value or "")
            values = []
            for col in range(2, max_col + 1):
                val = ws.cell(row=row_num, column=col).value
                if isinstance(val, (int, float)):
                    values.append(val)
            if values:
                result["expense_items"].append({"label": label, "values": values})
                result["total_expenses"] += sum(values) / max(len(values), 1)

        # Extract NOI
        noi_row = structure.get("noi_row")
        if noi_row:
            noi_values = []
            for col in range(2, max_col + 1):
                val = ws.cell(row=noi_row, column=col).value
                if isinstance(val, (int, float)):
                    noi_values.append(val)
            if noi_values:
                result["noi"] = noi_values[-1]  # use latest period
                result["noi_schedule"] = noi_values

        # Extract debt service
        ds_row = structure.get("debt_service_row")
        if ds_row:
            ds_values = []
            for col in range(2, max_col + 1):
                val = ws.cell(row=ds_row, column=col).value
                if isinstance(val, (int, float)):
                    ds_values.append(val)
            if ds_values:
                result["debt_service"] = ds_values[-1]

        # Extract occupancy
        occ_row = structure.get("occupancy_row")
        if occ_row:
            for col in range(2, max_col + 1):
                val = ws.cell(row=occ_row, column=col).value
                if isinstance(val, (int, float)):
                    pct = val if val > 1 else val * 100
                    result["occupancy_schedule"].append(round(pct, 1))

        return result

    def benchmark_assumptions(self, extracted: dict, asset_type: str = "senior_living") -> dict:
        """Compare extracted assumptions to NEST benchmarks.

        Returns per-assumption variance flags.
        """
        benchmarks = BENCHMARKS.get(asset_type, {})
        flags = []

        noi = extracted.get("noi")
        total_rev = extracted.get("total_revenue", 0)
        total_exp = extracted.get("total_expenses", 0)

        if noi and total_rev:
            margin = noi / total_rev if total_rev > 0 else 0
            if margin < 0.15:
                flags.append({
                    "metric": "noi_margin",
                    "value": round(margin * 100, 1),
                    "benchmark": "15-35%",
                    "flag": "aggressive" if margin < 0.10 else "low",
                    "note": "NOI margin below typical range",
                })

        if total_exp and total_rev:
            expense_ratio = total_exp / total_rev if total_rev > 0 else 1
            if expense_ratio > 0.75:
                flags.append({
                    "metric": "expense_ratio",
                    "value": round(expense_ratio * 100, 1),
                    "benchmark": "55-75%",
                    "flag": "high",
                    "note": "Operating expenses high relative to revenue",
                })

        occ = extracted.get("occupancy_schedule", [])
        if occ and asset_type == "senior_living":
            ramp = benchmarks.get("occupancy_ramp", {})
            for month_idx, (low, high) in ramp.items():
                if len(occ) >= month_idx:
                    actual = occ[month_idx - 1] / 100 if occ[month_idx - 1] > 1 else occ[month_idx - 1]
                    if actual > high:
                        flags.append({
                            "metric": f"occupancy_month_{month_idx}",
                            "value": round(actual * 100, 1),
                            "benchmark": f"{low*100:.0f}-{high*100:.0f}%",
                            "flag": "aggressive",
                            "note": f"Month {month_idx} occupancy above market benchmark",
                        })
                    elif actual < low:
                        flags.append({
                            "metric": f"occupancy_month_{month_idx}",
                            "value": round(actual * 100, 1),
                            "benchmark": f"{low*100:.0f}-{high*100:.0f}%",
                            "flag": "conservative",
                            "note": f"Month {month_idx} occupancy below market benchmark",
                        })

        return {
            "asset_type": asset_type,
            "flags": flags,
            "flags_count": len(flags),
            "aggressive_count": sum(1 for f in flags if f["flag"] == "aggressive"),
            "conservative_count": sum(1 for f in flags if f["flag"] == "conservative"),
        }

    def generate_ramp_model(self, assumptions: dict, units: int,
                             duration_months: int = 60) -> list:
        """Generate month-by-month occupancy + NOI + debt service schedule.

        Returns list of monthly snapshots.
        """
        monthly_rev_per_unit = assumptions.get("revenue_per_unit_monthly", 5200)
        annual_opex_per_unit = assumptions.get("opex_per_unit_annual", 55000)
        monthly_opex_per_unit = annual_opex_per_unit / 12
        debt_service_annual = assumptions.get("debt_service_annual", 0)
        monthly_ds = debt_service_annual / 12
        stabilization_month = assumptions.get("stabilization_month", 36)
        target_occupancy = assumptions.get("target_occupancy_pct", 90) / 100

        schedule = []
        cumulative_reserve_used = 0

        for month in range(1, duration_months + 1):
            # S-curve occupancy ramp
            if month <= stabilization_month:
                t = month / stabilization_month
                occ = target_occupancy * (3 * t**2 - 2 * t**3)  # smooth S-curve
            else:
                occ = target_occupancy

            occupied_units = int(units * occ)
            gross_revenue = occupied_units * monthly_rev_per_unit
            vacancy_loss = (units - occupied_units) * monthly_rev_per_unit
            total_opex = units * monthly_opex_per_unit * max(0.6, occ)  # some fixed costs
            noi = gross_revenue - total_opex
            dscr = noi / monthly_ds if monthly_ds > 0 else 0

            if noi < monthly_ds:
                deficit = monthly_ds - noi
                cumulative_reserve_used += deficit
            else:
                deficit = 0

            schedule.append({
                "month": month,
                "occupancy_pct": round(occ * 100, 1),
                "occupied_units": occupied_units,
                "gross_revenue": round(gross_revenue),
                "vacancy_loss": round(vacancy_loss),
                "total_opex": round(total_opex),
                "noi": round(noi),
                "debt_service": round(monthly_ds),
                "dscr": round(dscr, 3),
                "deficit": round(deficit),
                "cumulative_reserve_used": round(cumulative_reserve_used),
            })

        return schedule


# Singleton
proforma_spreader = ProformaSpreader()
